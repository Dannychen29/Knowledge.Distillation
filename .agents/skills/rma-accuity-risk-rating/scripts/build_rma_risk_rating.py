"""Build one traceable RMA risk-rating package from confirmed BA fields."""

from __future__ import annotations

import argparse
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook


SHEET_NAME = "風險表"
CELL_SOURCES = {
    "C3": ("Bank Name", "Roster"),
    "C5": ("BIC", "Roster"),
    "C6": ("Office Type", "Bankers Almanac (Accuity)"),
    "D12": ("World Rank", "Bankers Almanac (Accuity)"),
    "D14": ("Stock Symbol / Stock Exchange", "Bankers Almanac (Accuity)"),
    "D19": ("S&P Long Term", "Bankers Almanac (Accuity)"),
    "D20": ("Moody's Long Term", "Bankers Almanac (Accuity)"),
    "D21": ("Fitch Long Term", "Bankers Almanac (Accuity)"),
}


def formula_count(workbook):
    return sum(
        1
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )


def validation_count(workbook):
    return sum(len(sheet.data_validations.dataValidation) for sheet in workbook.worksheets)


def sha256(path):
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--bank-name", required=True)
    parser.add_argument("--bic", required=True)
    parser.add_argument("--ba-id", required=True)
    parser.add_argument("--office", choices=("Headquarter", "Branch"))
    parser.add_argument(
        "--office-source",
        default="Office Type",
        help="Visible BIC-matched Bankers Almanac office classification used for C6 traceability.",
    )
    parser.add_argument("--rank", type=int)
    parser.add_argument("--listed", choices=("Yes",))
    parser.add_argument("--sp")
    parser.add_argument("--moodys")
    parser.add_argument("--fitch")
    return parser.parse_args()


def main():
    args = parse_args()
    if not args.template.is_file():
        raise FileNotFoundError(args.template)
    args.output_dir.mkdir(parents=True, exist_ok=True)

    template = load_workbook(args.template, data_only=False)
    if SHEET_NAME not in template.sheetnames:
        raise ValueError(f"Missing worksheet: {SHEET_NAME}")
    original_sheet_names = template.sheetnames[:]
    original_formula_count = formula_count(template)
    original_validation_count = validation_count(template)
    sheet = template[SHEET_NAME]
    protected_labels = {cell: sheet[cell].value for cell in ("B3", "B4", "B5", "B6", "B7", "B8")}
    country_formula = sheet["C4"].value
    if not (isinstance(country_formula, str) and country_formula.startswith("=")):
        raise ValueError("C4 must contain the template country formula")

    values = {
        "C3": args.bank_name,
        "C5": args.bic,
        "C6": args.office,
        "D12": args.rank,
        "D14": args.listed,
        "D19": args.sp,
        "D20": args.moodys,
        "D21": args.fitch,
    }
    for cell, value in values.items():
        if value is not None:
            sheet[cell] = value

    workbook_path = args.output_dir / "filled-risk-rating.xlsx"
    template.save(workbook_path)
    with ZipFile(workbook_path) as archive:
        if archive.testzip() is not None:
            raise ValueError("Output workbook ZIP integrity check failed")

    verified = load_workbook(workbook_path, data_only=False)
    verified_sheet = verified[SHEET_NAME]
    if verified.sheetnames != original_sheet_names:
        raise ValueError("Worksheet names changed")
    if formula_count(verified) != original_formula_count:
        raise ValueError("Formula count changed")
    if validation_count(verified) != original_validation_count:
        raise ValueError("Data validation count changed")
    if verified_sheet["C4"].value != country_formula:
        raise ValueError("C4 formula changed")
    if any(verified_sheet[cell].value != value for cell, value in protected_labels.items()):
        raise ValueError("Protected label changed")
    for cell, value in values.items():
        if value is not None and verified_sheet[cell].value != value:
            raise ValueError(f"Confirmed mapping did not persist: {cell}")

    mappings = []
    for cell, value in values.items():
        source_field, source_system = CELL_SOURCES[cell]
        if cell == "C6":
            source_field = args.office_source
        status = "confirmed" if value is not None else "unavailable"
        mappings.append({
            "source_system": source_system,
            "source_field": source_field,
            "value": value,
            "target_sheet": SHEET_NAME,
            "target_cell": cell,
            "status": status,
            "bic": args.bic,
            "evidence_id": f"bankers-almanac-{args.bic}" if source_system.startswith("Bankers") else f"roster-{args.bic}",
            "ba_id": args.ba_id if source_system.startswith("Bankers") else None,
            "human_control": "manual review required",
        })
    (args.output_dir / "field-mapping.json").write_text(
        json.dumps({"mappings": mappings}, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    mapping_book = Workbook()
    mapping_sheet = mapping_book.active
    mapping_sheet.title = "Field Mapping"
    mapping_sheet.append(["Source system", "Source field", "Value", "Target sheet", "Target cell", "Status", "BIC", "BA ID", "Human control"])
    for mapping in mappings:
        mapping_sheet.append([
            mapping["source_system"], mapping["source_field"], mapping["value"],
            mapping["target_sheet"], mapping["target_cell"], mapping["status"],
            mapping["bic"], mapping["ba_id"], mapping["human_control"],
        ])
    mapping_book.save(args.output_dir / "field-mapping.xlsx")

    schema = {
        "template": str(args.template),
        "sheets": verified.sheetnames,
        "formula_count": original_formula_count,
        "validation_count": original_validation_count,
        "writable_cells": sorted(values),
    }
    (args.output_dir / "workbook-schema.json").write_text(
        json.dumps(schema, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    record = {
        "writer": "openpyxl",
        "writer_version": __import__("openpyxl").__version__,
        "template_sha256": sha256(args.template),
        "output_sha256": sha256(workbook_path),
        "ba_id": args.ba_id,
        "bic": args.bic,
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "verification": {
            "zip_ok": True,
            "sheets_preserved": True,
            "formulas_preserved": True,
            "validations_preserved": True,
            "country_formula_preserved": True,
            "protected_labels_preserved": True,
        },
    }
    (args.output_dir / "build-record.json").write_text(
        json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8"
    )


if __name__ == "__main__":
    main()
